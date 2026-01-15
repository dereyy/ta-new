import json
import networkx as nx
from typing import List, Set, Dict, Tuple
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import math
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class GLODAlgorithm:
    """
    Implementasi algoritma GLOD (Global Local Overlapping Community Detection)
    untuk deteksi komunitas overlapping pada jaringan protein-protein interaction.
    """
    
    def __init__(self, graph: nx.Graph, alpha: float = 0.8, jaccard_threshold: float = 0.33):
        self.graph = graph
        self.alpha = alpha
        self.jaccard_threshold = jaccard_threshold
        self.communities: List[Set] = []

    def omega(self, candidate: str, community: Set) -> float:
        """
        Hybrid node fitness function ω(vi) - Equation 4 from paper.
        
        Formula (dari paper Equation 4):
        ω(vi) = max_vj∈NCi [ (|N(vi)∩N(vj)|+1)/(|N(vj)|+1) + 0.1*(|N2(vi)∩N2(vj)|+1)/|N2(vj)| ] / D(vi)
        
        where:
        - N(vi) = neighbors of vi (orde 1)
        - N2(vi) = 2-hop neighbors of vi (orde 2)
        - NCi = N(vi) ∩ C (intersection of neighbors and community)
        - D(vi) = degree of candidate node vi (divisor yang benar)
        
        PENTING: Pembagi yang benar adalah D(vi), bukan 1.1.
        Normalisasi 1.1 hanya untuk menggabungkan bobot orde 1 dan 2,
        tetapi divisor luar kurung adalah derajat node vi.
        """
        neighbors_vi = set(self.graph.neighbors(candidate))
        NCi = neighbors_vi.intersection(community)
        if not NCi:
            return 0.0

        # 2-hop neighbors of vi
        N2_vi = set()
        for n in neighbors_vi:
            N2_vi.update(self.graph.neighbors(n))

        max_score = 0.0

        # Iterasi melalui semua node di NCi dan hitung similarity
        for vj in NCi:
            neighbors_vj = set(self.graph.neighbors(vj))

            # 2-hop neighbors of vj
            N2_vj = set()
            for n in neighbors_vj:
                N2_vj.update(self.graph.neighbors(n))

            # Penyebut orde 1: |N(vj)| + 1 sesuai Equation 4
            denom1 = len(neighbors_vj) + 1
            # Penyebut orde 2: |N2(vj)| (tanpa +1)
            denom2 = len(N2_vj) if N2_vj else 1

            # Orde 1 similarity: (|N(vi)∩N(vj)|+1) / (|N(vj)|+1)
            part1 = (len(neighbors_vi & neighbors_vj) + 1) / denom1
            # Orde 2 similarity dengan bobot 0.1: (|N2(vi)∩N2(vj)|+1) / |N2(vj)|
            part2 = (len(N2_vi & N2_vj) + 1) / denom2

            # Score menurut paper equation 4
            score = part1 + 0.1 * part2
            max_score = max(max_score, score)

        # PERBAIKAN: Pembagi harusnya D(vi) = derajat dari candidate node, bukan 1.1
        # Normalisasi 1.1 adalah untuk inner bracket, outer bracket dibagi D(vi)
        degree_vi = self.graph.degree(candidate)
        if degree_vi == 0:
            return 0.0
        
        return max_score / degree_vi
            
    def common_neighbor_similarity(self, node1, node2) -> int:
        """Hitung Common Neighbor Similarity (NC)"""
        neighbors1 = set(self.graph.neighbors(node1))
        neighbors2 = set(self.graph.neighbors(node2))
        return len(neighbors1.intersection(neighbors2))
    
    def calculate_seed_score(self, rough_seed: Set) -> float:
        """
        Hitung score untuk rough seed:
        Score = Sum(degree) + Count(Nodes) + Count(Internal Edges)
        """
        # Sum of degrees
        degree_sum = sum(self.graph.degree(node) for node in rough_seed)
        
        # Count nodes
        node_count = len(rough_seed)
        
        # Count internal edges
        edge_count = 0
        for node1 in rough_seed:
            for node2 in rough_seed:
                if node1 < node2 and self.graph.has_edge(node1, node2):
                    edge_count += 1
        
        return degree_sum + node_count + edge_count
    
    def create_rough_seed(self, center_node) -> Set:
        """
        Create rough seed Vi (Algorithm 1 dari paper).
        
        Sesuai Algorithm 1, line 8:
        Vi ← ({vi} ∪ N(vi))
        
        Proses:
        1. Mulai dengan center_node vi
        2. Tambahkan neighbors dengan NC (Common Neighbor Similarity) tertinggi
        3. Terus tambahkan neighbors dengan NC tertinggi sampai tidak ada lagi yang > 0
        4. Return rough seed Vi
        
        PENTING: Rough seed TIDAK diperluas pada tahap ini!
        Ekspansi dilakukan di tahap EXPANSION PHASE (Algorithm 2)
        """
        rough_seed = {center_node}
        neighbors = set(self.graph.neighbors(center_node))
        available_neighbors = neighbors - rough_seed
        
        # Fase konstruksi rough seed: tambahkan tetangga dengan NC tertinggi secara iteratif
        # Sesuai Algorithm 1, Line 8: Vi ← ({vi} ∪ N(vi))
        # Iterasi sampai tidak ada neighbor dengan NC > 0
        while available_neighbors:
            nc_scores = [
                (neighbor, self.common_neighbor_similarity(center_node, neighbor)) 
                for neighbor in available_neighbors
            ]
            # Sort by NC (descending), then by node ID (ascending) untuk tie-breaking konsisten
            nc_scores.sort(key=lambda x: (-x[1], x[0]))
            
            best_neighbor, best_nc = nc_scores[0]
            
            # Add neighbor jika NC > 0 (ada common neighbors)
            if best_nc > 0:
                rough_seed.add(best_neighbor)
                available_neighbors = neighbors - rough_seed
            else:
                # Tidak ada neighbor lagi dengan NC > 0, stop constructing rough seed
                break
        
        return rough_seed
    
    def fitness_function(self, community: Set) -> float:
        """
        Fitness function: f(C) = k_in / (k_in + k_out)^alpha
        """
        k_in = 0  # Internal degree
        k_out = 0  # External degree
        
        for node in community:
            for neighbor in self.graph.neighbors(node):
                if neighbor in community:
                    k_in += 1
                else:
                    k_out += 1
        
        if k_in == 0:
            return 0.0
        
        denominator = (k_in + k_out) ** self.alpha
        if denominator == 0:
            return 0.0
            
        return k_in / denominator
    
    def influence_function(self, candidate: str, seed: Set) -> float:
        """
        Influence function: F(v, S) = |N(v) ∩ S| / |S|
        """
        if len(seed) == 0:
            return 0.0
        
        candidate_neighbors = set(self.graph.neighbors(candidate))
        intersection = candidate_neighbors.intersection(seed)
        
        return len(intersection) / len(seed)
    
    def expand_seed(self, seed: Set) -> Set:
        """
        Expansion phase menggunakan OR logic (Algorithm 2 dari paper).
        
        Proses: Untuk setiap neighbor node dalam shell, hitung ketiga fungsi:
        - f(C) = fitness gain
        - ω(vi) = neighbor similarity
        - F(v,s) = influence
        
        Node ditambahkan jika ia merupakan argmax dari f(C) ATAU omega ATAU F(v,s).
        
        PENTING - PERBAIKAN STOPPING CONDITIONS:
        1. Fitness gain harus >= min_fitness_gain_threshold (0.0001 secara default)
           Ini adalah kriteria UTAMA untuk penambahan node
        
        2. Hanya jika fitness_gain negative NAMUN omega ATAU influence bernilai sangat tinggi (>0.8)
           Maka node boleh ditambahkan sebagai exception
        
        3. Jika tidak ada satupun argmax node yang memenuhi kriteria di atas,
           STOP ekspansi segera (break)
        
        4. Tambahan safeguard: jika community size melebihi threshold tertentu
           (misalnya 50% dari graph), stop untuk mencegah "giant component"
        """
        community = seed.copy()
        improved = True
        initial_fitness = self.fitness_function(community)
        iterations = 0
        min_fitness_gain_threshold = 0.0001  # Stopping condition yang ketat
        max_community_size_ratio = 0.5  # Jangan biarkan komunitas > 50% dari graph
        max_community_size = max(3, int(self.graph.number_of_nodes() * max_community_size_ratio))
        
        print(f"  Expanding seed (size: {len(seed)}, initial fitness: {initial_fitness:.4f})")
        print(f"  Max community size allowed: {max_community_size} nodes")
        
        while improved:
            improved = False
            current_fitness = self.fitness_function(community)
            
            # Dapatkan shell nodes (neighbors dari community yang belum di community)
            shell_nodes = set()
            for node in community:
                for neighbor in self.graph.neighbors(node):
                    if neighbor not in community:
                        shell_nodes.add(neighbor)
            
            if not shell_nodes:
                print(f"    Stopping: No shell nodes found")
                break
            
            # SAFEGUARD: Jika community sudah terlalu besar, stop
            if len(community) >= max_community_size:
                print(f"    Stopping: Community size ({len(community)}) reached max limit ({max_community_size})")
                break
            
            # Hitung semua nilai untuk semua candidates (Algorithm 2)
            # Sort shell_nodes untuk konsistensi iterasi
            candidate_scores = {}
            for candidate in sorted(shell_nodes):
                test_community = community.copy()
                test_community.add(candidate)
                
                # Hitung fitness gain
                fitness_gain = self.fitness_function(test_community) - current_fitness
                
                # Hitung omega
                omega_val = self.omega(candidate, community)
                
                # Hitung influence F(v,s)
                influence = self.influence_function(candidate, community)
                
                candidate_scores[candidate] = {
                    'fitness': fitness_gain,
                    'omega': omega_val,
                    'influence': influence
                }
            
            # Implementasi OR logic: cari node yang merupakan argmax dari salah satu fungsi
            # Langkah 1: Hitung argmax untuk setiap fungsi
            best_by_fitness = max(candidate_scores.items(), key=lambda x: x[1]['fitness'])
            best_by_omega = max(candidate_scores.items(), key=lambda x: x[1]['omega'])
            best_by_influence = max(candidate_scores.items(), key=lambda x: x[1]['influence'])
            
            # Langkah 2: Kumpulkan semua node yang merupakan argmax dari setidaknya satu fungsi
            # Gunakan set kemudian convert ke sorted list untuk konsistensi
            argmax_nodes_set = {
                best_by_fitness[0],
                best_by_omega[0],
                best_by_influence[0]
            }
            argmax_nodes = sorted(argmax_nodes_set)  # Sort untuk konsistensi iterasi
            
            # Langkah 3: Pilih node terbaik di antara argmax nodes dengan kriteria ketat
            best_candidate = None
            best_score = float('-inf')
            best_criterion = None
            best_fitness_gain = 0.0
            
            for candidate in argmax_nodes:
                scores = candidate_scores[candidate]
                
                # Untuk memilih node terbaik di antara argmax nodes, gunakan kombinasi score
                max_score_for_candidate = max(scores['fitness'], scores['omega'], scores['influence'])
                
                # Tie-breaking: jika score sama, pilih node ID terkecil
                if max_score_for_candidate > best_score or (
                    max_score_for_candidate == best_score and 
                    (best_candidate is None or candidate < best_candidate)
                ):
                    best_score = max_score_for_candidate
                    best_candidate = candidate
                    best_fitness_gain = scores['fitness']
                    if scores['fitness'] == max_score_for_candidate:
                        best_criterion = 'fitness'
                    elif scores['omega'] == max_score_for_candidate:
                        best_criterion = 'omega'
                    else:
                        best_criterion = 'influence'
            
            # KRITERIA STOPPING YANG KETAT:
            # 1. Kriteria utama: fitness_gain HARUS >= threshold
            # 2. Exception: omega atau influence SANGAT tinggi (>0.8) bisa override negatif fitness
            #    TAPI hanya jika tidak terlalu banyak iterasi sudah berjalan
            
            add_node = False
            reason = ""
            
            if best_candidate is not None:
                # PRIMARY CRITERION: fitness gain harus positif dan signifikan
                if best_fitness_gain >= min_fitness_gain_threshold:
                    add_node = True
                    reason = f"by {best_criterion} (fitness_gain: {best_fitness_gain:.6f})"
                
                # SECONDARY CRITERION: jika fitness negatif tapi omega/influence SANGAT tinggi
                # dan masih awal ekspansi (< 10 iterations)
                elif best_fitness_gain < 0 and iterations < 10:
                    scores = candidate_scores[best_candidate]
                    if scores['omega'] > 0.8:
                        add_node = True
                        reason = f"by omega={scores['omega']:.4f} (exceptional high similarity despite negative fitness)"
                    elif scores['influence'] > 0.8:
                        add_node = True
                        reason = f"by influence={scores['influence']:.4f} (exceptional high influence despite negative fitness)"
                
                # TERTIARY CRITERION: jika tidak memenuhi di atas, stop
                if not add_node and best_fitness_gain < min_fitness_gain_threshold:
                    print(f"    Stopping expansion: No candidate meets criteria")
                    print(f"      Best: {best_candidate} by {best_criterion}")
                    print(f"      fitness_gain: {best_fitness_gain:.6f} (threshold: {min_fitness_gain_threshold})")
                    print(f"      omega: {candidate_scores[best_candidate]['omega']:.4f}")
                    print(f"      influence: {candidate_scores[best_candidate]['influence']:.4f}")
                    break
            
            if add_node and best_candidate is not None:
                community.add(best_candidate)
                improved = True
                iterations += 1
                print(f"    Added node {best_candidate} {reason} (community size: {len(community)})")
            else:
                break
        
        final_fitness = self.fitness_function(community)
        print(f"    Expansion complete: {len(seed)} → {len(community)} nodes, fitness: {initial_fitness:.4f} → {final_fitness:.4f} ({iterations} iterations)")
        return community
    
    def jaccard_coefficient(self, comm1: Set, comm2: Set) -> float:
        """Hitung Jaccard Coefficient standar untuk dua komunitas"""
        intersection = len(comm1.intersection(comm2))
        union = len(comm1.union(comm2))
        
        if union == 0:
            return 0.0
        
        return intersection / union
    
    def improved_jaccard_coefficient(self, comm_idx1: int, comm_idx2: int) -> float:
        """
        Improved Jaccard Coefficient (Equation 8 dari paper) untuk overlapping nodes.
        
        Formula Paper (Equation 8):
        J(o_vi) = Σ_{ovi∈(Ci∩Cj)} (1/|Ci ∪ Cj|)
        
        Artinya: Untuk setiap node overlapping (node yang ada di KEDUA komunitas),
        hitung kontribusinya sebagai 1/|Ci ∪ Cj|.
        
        Kemudian hitung total bobot overlapping:
        J_total = Σ(kontribusi per overlapping node)
        
        Jika J_total >= jaccard_threshold (default 1/3), merge komunitas tersebut.
        """
        comm1 = self.communities[comm_idx1]
        comm2 = self.communities[comm_idx2]
        
        # Temukan overlapping nodes (node yang ada di KEDUA komunitas)
        overlapping_nodes = comm1.intersection(comm2)
        
        if not overlapping_nodes:
            return 0.0
        
        # Hitung union dari kedua komunitas
        union_size = len(comm1.union(comm2))
        if union_size == 0:
            return 0.0
        
        # Equation 8: J(o_vi) = Σ (1/|Ci ∪ Cj|) untuk setiap overlapping node
        # Karena semua overlapping node memiliki kontribusi yang sama (1/union_size),
        # maka total = |overlapping_nodes| * (1/union_size)
        # = |overlapping_nodes| / union_size
        improved_jaccard = len(overlapping_nodes) / union_size
        
        return improved_jaccard
    
    def merge_communities(self):
        """
        Merge phase (Algorithm 3 dari paper).
        
        Untuk setiap pair komunitas yang memiliki overlapping nodes:
        - Hitung improved Jaccard coefficient J = Σ 1/|Ci ∪ Cj| (Equation 8)
        - Jika J >= jaccard_threshold (user input), merge komunitas tersebut
        
        PENTING: Paper menggunakan 1/3 sebagai default, tapi kami memungkinkan user 
        untuk mengatur threshold ini secara dinamis untuk eksperimentasi.
        """
        print(f"\nStarting merge phase with improved Jaccard coefficient (threshold: {self.jaccard_threshold:.4f})")
        print(f"Communities before merging: {len(self.communities)}")
        
        merge_count = 0
        merged = True
        
        while merged:
            merged = False
            new_communities = []
            communities_to_skip = set()
            
            for i in range(len(self.communities)):
                if i in communities_to_skip:
                    continue
                
                current_comm = self.communities[i]
                communities_to_merge = [i]  # Mulai dengan komunitas saat ini
                
                # Cari semua overlapping dengan komunitas lain
                for j in range(i + 1, len(self.communities)):
                    if j in communities_to_skip:
                        continue
                    
                    other_comm = self.communities[j]
                    
                    # Hitung improved Jaccard coefficient (Equation 8)
                    improved_jaccard = self.improved_jaccard_coefficient(i, j)
                    
                    # Jika J >= threshold, mark untuk merge
                    if improved_jaccard >= self.jaccard_threshold:
                        print(f"  Merging C{i} (size={len(current_comm)}) and C{j} (size={len(other_comm)}) " +
                              f"(Improved Jaccard: {improved_jaccard:.4f} >= {self.jaccard_threshold:.4f})")
                        communities_to_merge.append(j)
                        communities_to_skip.add(j)
                        merged = True
                        merge_count += 1
                
                # Merge semua komunitas yang dipilih
                if len(communities_to_merge) > 1:
                    merged_community = current_comm.copy()
                    for idx in communities_to_merge[1:]:
                        merged_community = merged_community.union(self.communities[idx])
                    new_communities.append(merged_community)
                else:
                    new_communities.append(current_comm)
                
                communities_to_skip.add(i)
            
            self.communities = new_communities
        
        print(f"Merging complete: {merge_count} merges performed")
        print(f"Communities after merging: {len(self.communities)}\n")
    
    def calculate_shen_modularity(self) -> float:
        """
        Menghitung Extended Modularity (EQ) berdasarkan paper Shen et al.
        EQ mengevaluasi kebaikan cover (overlapping community structure).
        
        Formula Persamaan (2) Shen et al.:
        EQ = (1/2m) * Σ_i Σ_j Σ_c (1/(O_v * O_w)) * (A_vw - k_v*k_w/2m)
        
        Dimana:
        - O_v = jumlah komunitas yang mengandung vertex v
        - A_vw = 1 jika ada edge (v,w), 0 jika tidak
        - k_v = degree dari vertex v
        - m = total edges dalam graph
        """
        m = self.graph.number_of_edges()
        if m == 0:
            return 0.0
        
        # Persiapan variabel: O_v adalah jumlah komunitas yang diikuti oleh vertex v
        Ov = {}
        for node in self.graph.nodes():
            count = sum(1 for comm in self.communities if node in comm)
            Ov[node] = count if count > 0 else 1  # Hindari pembagian dengan nol
        
        total_eq = 0.0
        
        # Implementasi Persamaan (2) Shen et al.
        for community in self.communities:
            nodes_in_comm = list(community)
            for i in range(len(nodes_in_comm)):
                node_v = nodes_in_comm[i]
                k_v = self.graph.degree(node_v)
                o_v = Ov[node_v]
                
                for j in range(len(nodes_in_comm)):
                    node_w = nodes_in_comm[j]
                    k_w = self.graph.degree(node_w)
                    o_w = Ov[node_w]
                    
                    # A_vw: 1 jika bertetangga, 0 jika tidak
                    a_vw = 1 if self.graph.has_edge(node_v, node_w) else 0
                    
                    # Suku dalam kurung: [A_vw - (k_v * k_w / 2m)]
                    expected_edges = (k_v * k_w) / (2 * m)
                    actual_minus_expected = a_vw - expected_edges
                    
                    # Tambahkan kontribusi ke EQ dengan bobot 1/(O_v * O_w)
                    total_eq += (1.0 / (o_v * o_w)) * actual_minus_expected
        
        # Normalisasi akhir dengan 1/(2m)
        return total_eq / (2 * m)
    
    def calculate_lazar_modularity(self) -> float:
        """
        Implementasi Lázár Modularity (M^ov) - Persamaan 4 & 5 dari paper Lázár et al.
        
        Metrik ini mengevaluasi keseimbangan antara kepadatan internal komunitas 
        dan rasio edge internal vs eksternal per node.
        
        Formula:
        M^ov_cr = (1/n_cr) * Σ_i∈cr [(k_in,i - k_out,i) / (d_i * s_i)] * density_cr
        
        Dimana:
        - n_cr = jumlah node dalam komunitas r
        - k_in,i = jumlah neighbor i yang ada di komunitas r
        - k_out,i = jumlah neighbor i yang di luar komunitas r
        - d_i = derajat total node i
        - s_i = jumlah komunitas yang mengandung node i
        - density_cr = n_e,cr / (n_cr choose 2) = kepadatan internal komunitas
        
        Nilai akhir: M^ov = (1/K) * Σ_r M^ov_cr (rata-rata dari semua komunitas)
        """
        K = len(self.communities)
        if K == 0:
            return 0.0
        
        # Hitung s_i: jumlah komunitas yang mengandung setiap node
        si = {}
        for node in self.graph.nodes():
            count = sum(1 for c in self.communities if node in c)
            si[node] = count if count > 0 else 1
        
        total_lazar = 0.0
        
        for community in self.communities:
            n_cr = len(community)
            
            # Skip komunitas dengan ukuran < 2 (tidak ada edge internal yang mungkin)
            if n_cr < 2:
                continue
            
            # Hitung densitas internal: edge internal / max possible edges
            internal_edges = self.graph.subgraph(community).number_of_edges()
            max_possible_edges = (n_cr * (n_cr - 1)) / 2
            density = internal_edges / max_possible_edges if max_possible_edges > 0 else 0.0
            
            # Hitung kontribusi node dalam komunitas
            node_contributions = 0.0
            for i in community:
                neighbors_i = set(self.graph.neighbors(i))
                
                # k_in: neighbors yang ada di komunitas
                k_in = len(neighbors_i.intersection(community))
                
                # k_out: neighbors yang di luar komunitas
                k_out = len(neighbors_i) - k_in
                
                # d_i: total degree
                d_i = self.graph.degree(i)
                
                # Sesuai Persamaan 2 Lázár: (k_in - k_out) / (d_i * s_i)
                if d_i > 0:
                    node_contributions += (k_in - k_out) / (d_i * si[i])
            
            # M^ov_cr = (average contribution per node) * density
            m_cr = (node_contributions / n_cr) * density
            total_lazar += m_cr
        
        # Rata-rata dari semua komunitas
        return total_lazar / K
    
    def calculate_nicosia_modularity(self) -> float:
        """
        Implementasi Nicosia Modularity (Q_ov) - Persamaan 15 adaptasi untuk Undirected Graph
        
        Metrik ini menggunakan "belonging factor" untuk mengakomodasi node yang berada 
        di banyak komunitas, dan memperluas modularitas Newman ke kasus overlapping.
        
        Formula (undirected adaptation):
        Q_ov = (1/2m) * Σ_c Σ_{i,j∈c} [β_ij,c * A_ij - β_ij,c * (k_i * k_j / 2m)]
        
        Dimana:
        - β_ij,c = α_i,c * α_j,c (belonging factor untuk edge di komunitas c)
        - α_i,c = 1/s_i jika node i ∈ c, 0 jika tidak (non-fuzzy assumption)
        - s_i = jumlah komunitas yang mengandung node i
        - A_ij = adjacency matrix (1 jika edge ada, 0 jika tidak)
        - k_i = degree dari node i
        - m = total edges dalam graph
        
        Note: Menggunakan produk belonging factor sesuai saran Nicosia Section 3.1
        """
        m = self.graph.number_of_edges()
        if m == 0:
            return 0.0
        
        # Hitung s_i: jumlah komunitas yang mengandung setiap node
        si = {}
        for node in self.graph.nodes():
            count = sum(1 for c in self.communities if node in c)
            si[node] = count if count > 0 else 1
        
        q_ov = 0.0
        
        # Iterasi per komunitas
        for community in self.communities:
            nodes_in_comm = list(community)
            
            # Iterasi semua pair node dalam komunitas
            for idx_i in range(len(nodes_in_comm)):
                node_i = nodes_in_comm[idx_i]
                k_i = self.graph.degree(node_i)
                
                for idx_j in range(len(nodes_in_comm)):
                    node_j = nodes_in_comm[idx_j]
                    k_j = self.graph.degree(node_j)
                    
                    # Belonging factor: β_ij,c = α_i,c * α_j,c
                    # Dimana α_i,c = 1/s_i untuk non-fuzzy case
                    beta_ij = (1.0 / si[node_i]) * (1.0 / si[node_j])
                    
                    # Adjacency matrix: 1 jika edge ada, 0 jika tidak
                    aij = 1 if self.graph.has_edge(node_i, node_j) else 0
                    
                    # Null model Newman (undirected): k_i * k_j / 2m
                    null_model = (k_i * k_j) / (2 * m)
                    
                    # Kontribusi modularity: β * (A - null_model)
                    q_ov += beta_ij * (aij - null_model)
        
        # Normalisasi dengan 2m
        return q_ov / (2 * m)
    
    # Alias untuk backward compatibility
    def calculate_modularity(self) -> float:
        """Backward compatibility - menggunakan Shen modularity sebagai default"""
        return self.calculate_shen_modularity()
    def calculate_psi_normalized_node_cut(self, community: Set) -> float:
        """
        Menghitung Normalized Node Cut (Psi) berdasarkan Havemann et al. (2012).
        Formula: Psi = (1/k_in) * Σ [ (k_in_i * k_out_i) / k_i ]
        """
        if not community:
            return 0.0
        
        # 1. Hitung total internal degree komunitas (k_in)
        # Di paper, k_in(C) adalah jumlah degree dari semua node internal [cite: 4105, 4403]
        total_k_in_community = 0
        
        # 2. Iterasi setiap node untuk menghitung sigma
        psi_numerator_sum = 0.0
        
        for node in community:
            if node not in self.graph:
                continue
                
            k_i_in = 0   # internal degree node i
            k_i_out = 0  # external degree node i
            
            for neighbor in self.graph[node]:
                weight = self.graph[node][neighbor].get('weight', 1)
                if neighbor in community:
                    k_i_in += weight
                else:
                    k_i_out += weight
            
            # k_i adalah total degree node tersebut [cite: 4098]
            k_i = k_i_in + k_i_out
            
            if k_i > 0:
                # Formula inti: (k_in_i * k_out_i) / k_i 
                psi_numerator_sum += (k_i_in * k_i_out) / k_i
                
            # Tambahkan ke total k_in komunitas
            total_k_in_community += k_i_in

        if total_k_in_community == 0:
            return 0.0
            
        # Hasil akhir adalah sumasi dibagi total internal degree 
        return psi_numerator_sum / total_k_in_community
    
    def calculate_conductance(self, community: Set) -> float:
        """
        Menghitung Conductance untuk komunitas tunggal.
        
        Implementasi berdasarkan Leskovec et al.
        Formula: Conductance = k_out / (k_in + k_out)
        
        Dimana:
        - k_out: jumlah edge yang meninggalkan komunitas (cut edges)
        - k_in: jumlah edge internal dalam komunitas
        
        Interpretasi: Semakin rendah nilainya, semakin baik kualitas boundary
        komunitas (Lower is better - mengukur kualitas batas komunitas)
        
        Args:
            community: Set dari node IDs yang membentuk komunitas
            
        Returns:
            float: Nilai Conductance untuk komunitas (0 jika komunitas kosong)
        """
        if not community:
            return 0.0
        
        # Hitung jumlah edge internal (k_in)
        k_in = 0
        for node1 in community:
            for node2 in community:
                if node1 < node2:
                    if self.graph.has_edge(node1, node2):
                        k_in += self.graph[node1][node2].get('weight', 1)
        
        # Hitung jumlah edge eksternal (cut edges) (k_out)
        k_out = 0
        for node in community:
            if node in self.graph:
                for neighbor in self.graph[node]:
                    if neighbor not in community:
                        k_out += self.graph[node][neighbor].get('weight', 1)
        
        # Jika tidak ada edge sama sekali, return 0
        if k_in + k_out == 0:
            return 0.0
        
        return k_out / (k_in + k_out)
    
    def h_binary(self, w: int, n: int) -> float:
        """
        Fungsi entropi biner sesuai Equation 1 dari McDaid et al.
        h(w, n) = -[w/n * log2(w/n) + (n-w)/n * log2((n-w)/n)]
        
        Jika w = 0 atau w = n, maka h = 0 (entropi minimal)
        """
        if w == 0 or w == n:
            return 0.0
        
        p = w / n
        q = (n - w) / n
        
        # Hindari log(0)
        term1 = p * math.log2(p) if p > 0 else 0
        term2 = q * math.log2(q) if q > 0 else 0
        
        return -(term1 + term2)
    
    def get_entropy_single(self, comm: Set, n: int) -> float:
        """
        Menghitung entropi dari satu komunitas.
        H(C_i) = h(|C_i|, n) + h(n - |C_i|, n)
        
        Dimana h adalah fungsi entropi biner.
        """
        w = len(comm)
        return self.h_binary(w, n) + self.h_binary(n - w, n)
    
    def get_conditional_entropy_optimized(self, X_comms: List[Set], Y_comms: List[Set], n: int) -> float:
        """
        Menghitung H(X|Y) = Σ_i min_j H*(X_i|Y_j)
        
        Sesuai dengan Equation B.14 dari Lancichinetti et al dan Equation 2 dari McDaid et al.
        
        Untuk setiap partisi X_i dalam X, cari partisi Y_j dalam Y yang memiliki 
        conditional entropy minimum, dengan constraint matching.
        """
        total_h_x_given_y = 0.0
        
        for xi in X_comms:
            h_xi = self.get_entropy_single(xi, n)
            min_h_xi_yj = h_xi  # Default: jika tidak ada matching yang lebih baik
            
            for yj in Y_comms:
                # Buat tabel kontingensi 2x2:
                # a = not in xi, not in yj
                # b = not in xi, in yj
                # c = in xi, not in yj
                # d = in xi, in yj
                
                d = len(xi & yj)  # Intersection
                c = len(xi) - d   # In xi but not yj
                b = len(yj) - d   # In yj but not xi
                a = n - (b + d + c)  # Not in either
                
                # Hitung H* sesuai Equation 2 McDaid: 
                # H*(X_i|Y_j) = H(a,b) + H(c,d) - H(a+c) - H(b+d)
                # dimana H(x,y) = h_binary(x, x+y)
                
                term1 = self.h_binary(a, a + b) if (a + b) > 0 else 0
                term2 = self.h_binary(c, c + d) if (c + d) > 0 else 0
                term3 = self.h_binary(a + c, n)
                term4 = self.h_binary(b + d, n)
                
                h_xi_yj = term1 + term2 - term3 - term4
                
                # Cari minimum
                if h_xi_yj < min_h_xi_yj:
                    min_h_xi_yj = h_xi_yj
            
            total_h_x_given_y += min_h_xi_yj
        
        return total_h_x_given_y
    
    def calculate_onmi_metrics(self, detected_comms: List[Set], ground_truth_comms: List[Set], seed_value: int = 42) -> Dict[str, float]:
        """
        Menghitung tiga metrik NMI untuk evaluasi overlapping community detection.
        
        Parameters:
        - detected_comms: List dari Set, hasil deteksi komunitas dari algoritma
        - ground_truth_comms: List dari Set, label ground truth komunitas yang sebenarnya
        - seed_value: Random seed untuk konsistensi perhitungan rNMI (default: 42)
        
        Returns:
        Dictionary dengan keys: 'nmi_lfk', 'nmi_max', 'rnmi'
        
        ===== PENJELASAN KETIGA METRIK =====
        
        1. NMI_LFK (Lancichinetti, Fortunato, Kertész):
           Formula: NMI_LFK = 1 - 0.5 * [H(X|Y)/H(X) + H(Y|X)/H(Y)]
           Normalisasi berdasarkan entropi kondisional rata-rata.
           Range: [0, 1], dimana 1 = perfect match
        
        2. NMI_max (McDaid, Greene, Hurley) - perbaikan dari LFK:
           Formula: NMI_max = I(X:Y) / max(H(X), H(Y))
           Normalisasi menggunakan entropi maksimum (lebih konvensional).
           Lebih stabil dan interpretable dibanding LFK.
           Range: [0, 1], dimana 1 = perfect match
        
        3. rNMI (Relative NMI, Pan Zhang) - mengatasi finite-size effect:
           Formula: rNMI = NMI_max - E[NMI_max(A, random_B)]
           Mengurangi nilai dengan ekspektasi NMI dari partisi acak.
           Nilai 0 benar-benar berarti "tidak lebih baik dari acak".
           Range: [-1, 1], dimana nilai negatif = "lebih buruk dari acak"
        """
        num_nodes = len(self.graph.nodes())
        
        if num_nodes == 0 or not detected_comms or not ground_truth_comms:
            return {
                'nmi_lfk': 0.0,
                'nmi_max': 0.0,
                'rnmi': 0.0
            }
        
        # 1. Hitung Entropi Total H(X) dan H(Y)
        hx = sum(self.get_entropy_single(c, num_nodes) for c in detected_comms)
        hy = sum(self.get_entropy_single(c, num_nodes) for c in ground_truth_comms)
        
        # 2. Hitung Conditional Entropy H(X|Y) dan H(Y|X)
        hx_y = self.get_conditional_entropy_optimized(detected_comms, ground_truth_comms, num_nodes)
        hy_x = self.get_conditional_entropy_optimized(ground_truth_comms, detected_comms, num_nodes)
        
        # 3. Hitung Mutual Information I(X:Y) sesuai McDaid Equation 5:
        # I(X:Y) = 0.5 * [(H(X) - H(X|Y)) + (H(Y) - H(Y|X))]
        mutual_info = 0.5 * ((hx - hx_y) + (hy - hy_x))
        
        # ===== A. NMI_LFK (Lancichinetti) =====
        # Normalisasi LFK: 1 - 0.5*(H(X|Y)/H(X) + H(Y|X)/H(Y))
        if hx > 0 and hy > 0:
            term1 = hx_y / hx
            term2 = hy_x / hy
            nmi_lfk = 1.0 - 0.5 * (term1 + term2)
        else:
            nmi_lfk = 1.0 if hx == 0 and hy == 0 else 0.0
        
        # ===== B. NMI_max (McDaid / MGH) =====
        # Normalisasi konvensional: I(X:Y) / max(H(X), H(Y))
        max_entropy = max(hx, hy)
        if max_entropy > 0:
            nmi_max = mutual_info / max_entropy
        else:
            nmi_max = 1.0 if mutual_info == 0 else 0.0
        
        # ===== C. rNMI (Pan Zhang) - Relative NMI =====
        # rNMI = NMI_max - E[NMI_max(detected, random_ground_truth)]
        # Buat beberapa ground truth acak dan hitung rata-rata NMI-nya
        
        import random
        random.seed(seed_value)  # Set seed untuk konsistensi hasil
        random_nmis = []
        nodes_list = sorted(list(self.graph.nodes()))  # Sort untuk konsistensi
        
        # Generate 10 partisi acak dengan distribusi ukuran yang sama
        for _ in range(10):
            random_comms = []
            for gt_c in ground_truth_comms:
                # Buat partisi acak dengan ukuran yang sama dengan komunitas ground truth
                size = len(gt_c)
                random_comm = set(random.sample(nodes_list, min(size, num_nodes)))
                random_comms.append(random_comm)
            
            # Hitung NMI_max untuk pasangan (detected_comms, random_comms)
            sh_rand = sum(self.get_entropy_single(c, num_nodes) for c in random_comms)
            hx_rand = self.get_conditional_entropy_optimized(detected_comms, random_comms, num_nodes)
            hy_rand = self.get_conditional_entropy_optimized(random_comms, detected_comms, num_nodes)
            mi_rand = 0.5 * ((hx - hx_rand) + (sh_rand - hy_rand))
            
            max_ent = max(hx, sh_rand)
            if max_ent > 0:
                nmi_rand = mi_rand / max_ent
            else:
                nmi_rand = 0.0
            
            random_nmis.append(nmi_rand)
        
        # rNMI = NMI_max - E[NMI_random]
        avg_random_nmi = sum(random_nmis) / len(random_nmis) if random_nmis else 0.0
        rnmi = nmi_max - avg_random_nmi
        
        print(f"\n{'='*60}")
        print(f"ONMI Metrics Calculation:")
        print(f"  H(X) [detected]:     {hx:.4f}")
        print(f"  H(Y) [ground_truth]: {hy:.4f}")
        print(f"  H(X|Y):              {hx_y:.4f}")
        print(f"  H(Y|X):              {hy_x:.4f}")
        print(f"  I(X:Y):              {mutual_info:.4f}")
        print(f"  Avg Random NMI:      {avg_random_nmi:.4f}")
        print(f"{'='*60}")
        print(f"NMI Results:")
        print(f"  NMI_LFK (Lancichinetti):  {max(0, nmi_lfk):.4f}")
        print(f"  NMI_max (McDaid):         {max(0, nmi_max):.4f}")
        print(f"  rNMI (Relative):          {rnmi:.4f}")
        print(f"{'='*60}\n")
        
        return {
            'nmi_lfk': max(0.0, nmi_lfk),      # Clamp ke [0, 1]
            'nmi_max': max(0.0, nmi_max),      # Clamp ke [0, 1]
            'rnmi': rnmi                        # Bisa negatif
        }
    
    def run(self, seed_value: int = 42) -> Tuple[List[Set], float, float, float]:
        """
        Jalankan algoritma GLOD lengkap sesuai dengan paper.
        
        Tiga fase utama (sesuai Algorithm 1, 2, 3 dari paper):
        1. Seeding Phase (Algorithm 1): Pilih seed dari nodes dalam NL (Nodes without Labels)
                                        NL adalah koleksi node yang BELUM memiliki komunitas
                                        Setelah node masuk komunitas, HAPUS dari NL agar tidak diproses ulang
        2. Expansion Phase (Algorithm 2): Ekspansi seed dengan OR logic (fitness, omega, influence)
        3. Merge Phase (Algorithm 3): Merge komunitas overlapping dengan improved Jaccard >= 1/3
        
        Args:
            seed_value: Random seed untuk reproducible results (default: 42)
        
        Returns: (list of communities, shen_modularity, lazar_modularity, nicosia_modularity)
        
        PERBAIKAN KRITIS: 
        - Implementasi NL yang benar: Node yang sudah masuk komunitas DIHAPUS dari NL
        - Ini mencegah redundansi tinggi dan mempercepat eksekusi drastis
        - Overlapping TETAP BISA terjadi namun dengan seed yang berbeda-beda
        - SET RANDOM SEED untuk reproducibility: hasil akan konsisten dengan parameter yang sama
        """
        import random
        random.seed(seed_value)
        
        print(f"Starting GLOD with {self.graph.number_of_nodes()} nodes and {self.graph.number_of_edges()} edges")
        print(f"Random seed set to {seed_value} for reproducible results")
        
        # Phase 1: Seeding (Algorithm 1)
        # NL = koleksi node yang belum memiliki label komunitas (dapat diproses sebagai seed)
        # Berbeda dengan NC (node yang sudah menjadi center), NL adalah nodes yang ELIGIBLE untuk seeding
        NL = set(self.graph.nodes())  # All nodes start as unlabeled
        all_nodes = set(self.graph.nodes())
        candidate_seeds = []
        
        print(f"Seeding phase: Starting with {len(NL)} unlabeled nodes")
        
        iteration = 0
        max_iterations = 1000  # Safety limit
        
        # Sesuai Algorithm 1: while NL != empty, process nodes
        while NL and iteration < max_iterations:
            iteration += 1
            
            # Pilih node dengan derajat tertinggi dari NL sebagai center (Algorithm 1, line 3)
            # Sorting by centrality ensures good seed selection
            # PENTING: Gunakan sorting dengan tie-breaking konsisten menggunakan node ID
            best_center = min(
                NL, 
                key=lambda node: (-self.graph.degree(node), node)  # Negative degree untuk max, node ID untuk tie-break
            )
            
            print(f"\nIteration {iteration}: Processing center node {best_center} (degree: {self.graph.degree(best_center)}, NL size: {len(NL)})")
            
            # Buat rough seed dari center node (Algorithm 1, line 8: Vi = {vi} ∪ N(vi))
            rough_seed = self.create_rough_seed(best_center)
            score = self.calculate_seed_score(rough_seed)
            
            print(f"  Rough seed created: size {len(rough_seed)}, score {score:.2f}")
            
            candidate_seeds.append((rough_seed, score, best_center))
            
            # PENTING: Hapus center node dari NL (Algorithm 1, line 2)
            # Ini mengikuti logika paper bahwa node yang sudah diproses dihapus dari kandidat
            NL.discard(best_center)
            
            # Safety break: jika NL terlalu besar, hanya proses top candidates
            if len(candidate_seeds) >= 100:
                print(f"  Reached 100 candidate seeds, processing them now...")
                break
        
        # Urutkan berdasarkan score untuk prioritas ekspansi
        candidate_seeds.sort(key=lambda x: x[1], reverse=True)
        print(f"\nSeeding phase complete: {len(candidate_seeds)} candidate seeds found")
        
        # Phase 2: Expansion (Algorithm 2)
        # Proses candidate seeds dan ekspansi ke komunitas
        processed_seeds = set()
        nodes_in_communities = set()  # Track which nodes have been assigned
        
        print(f"\nExpansion phase: Processing {len(candidate_seeds)} candidate seeds")
        
        for seed_idx, (candidate_seed, score, center_node) in enumerate(candidate_seeds, 1):
            # Hindari seed duplicate
            seed_tuple = tuple(sorted(candidate_seed))
            if seed_tuple in processed_seeds:
                continue
            
            processed_seeds.add(seed_tuple)
            
            print(f"\n  Seed {seed_idx}/{len(candidate_seeds)}: center={center_node}, seed_size={len(candidate_seed)}")
            
            # Ekspansi seed (Algorithm 2)
            community = self.expand_seed(candidate_seed)
            
            # Simpan komunitas jika valid (minimal 3 node)
            if len(community) >= 3:
                self.communities.append(community)
                nodes_in_communities.update(community)
                print(f"    Community saved: {len(community)} nodes")
            else:
                print(f"    Community rejected: size {len(community)} < 3")
        
        # Hitung statistics
        unlabeled_nodes = all_nodes - nodes_in_communities
        
        print(f"\n{'='*60}")
        print(f"After expansion phase:")
        print(f"  Total communities: {len(self.communities)}")
        print(f"  Nodes in at least one community: {len(nodes_in_communities)}/{len(all_nodes)}")
        print(f"  Unlabeled nodes: {len(unlabeled_nodes)}")
        
        # Hitung overlapping statistics
        node_community_count = {}
        for node in nodes_in_communities:
            count = sum(1 for comm in self.communities if node in comm)
            node_community_count[node] = count
        
        overlapping_nodes = sum(1 for count in node_community_count.values() if count > 1)
        print(f"  Overlapping nodes (in multiple communities): {overlapping_nodes}")
        print(f"{'='*60}\n")
        
        # Phase 3: Merging (Algorithm 3)
        self.merge_communities()
        
        print(f"Final {len(self.communities)} communities after merging")
        print(f"Parameters used: alpha={self.alpha}, jaccard_threshold={self.jaccard_threshold}")
        
        # Hitung semua modularitas metrics
        print(f"\nCalculating modularity metrics...")
        shen_eq = self.calculate_shen_modularity()
        lazar_mov = self.calculate_lazar_modularity()
        nicosia_qov = self.calculate_nicosia_modularity()
        
        print(f"Calculated Shen Modularity (EQ): {shen_eq:.4f}")
        print(f"Calculated Lázár Modularity (M^ov): {lazar_mov:.4f}")
        print(f"Calculated Nicosia Modularity (Q_ov): {nicosia_qov:.4f}\n")
        
        return self.communities, shen_eq, lazar_mov, nicosia_qov


@require_http_methods(["GET", "POST"])
def glod_process(request):
    """Halaman input parameter GLOD dan proses algoritma"""
    
    print(f"glod_process called with method: {request.method}")
    
    if request.method == "POST":
        # Ambil data jaringan dari POST
        network_data_json = request.POST.get('network_data')
        
        print(f"Received network_data_json: {network_data_json[:100] if network_data_json else 'None'}...")
        
        if not network_data_json:
            return render(request, 'glod_app/process.html', {
                'error': 'Tidak ada data jaringan yang diterima'
            })
        
        try:
            network_data = json.loads(network_data_json)
            print(f"Parsed network data: {len(network_data.get('nodes', []))} nodes, {len(network_data.get('edges', []))} edges")
        except json.JSONDecodeError as e:
            print(f"JSON decode error: {e}")
            return render(request, 'glod_app/process.html', {
                'error': f'Format data jaringan tidak valid: {str(e)}'
            })
        
        # Simpan ke session untuk diproses
        request.session['glod_network_data'] = network_data
        request.session.save()
        
        print("Network data saved to session")
        
        # Tampilkan halaman input parameter
        context = {
            'total_nodes': len(network_data.get('nodes', [])),
            'total_edges': len(network_data.get('edges', [])),
            'default_alpha': 0.8,
            'default_threshold': 0.33
        }
        
        return render(request, 'glod_app/process.html', context)
    
    # GET request - cek apakah ada data di session
    network_data = request.session.get('glod_network_data')
    
    if not network_data:
        return render(request, 'glod_app/process.html', {
            'error': 'Tidak ada data jaringan. Silakan build network terlebih dahulu dari halaman String Network.'
        })
    
    context = {
        'total_nodes': len(network_data.get('nodes', [])),
        'total_edges': len(network_data.get('edges', [])),
        'default_alpha': 0.8,
        'default_threshold': 0.33
    }
    
    return render(request, 'glod_app/process.html', context)


@require_http_methods(["POST"])
def glod_result(request):
    """Proses algoritma GLOD dan tampilkan hasil"""
    
    # Ambil parameter dari form
    alpha = float(request.POST.get('alpha', 0.8))
    jaccard_threshold = float(request.POST.get('jaccard_threshold', 0.33))
    
    print(f"\n{'='*60}")
    print(f"GLOD Algorithm Parameters:")
    print(f"Alpha (α): {alpha}")
    print(f"Jaccard Threshold: {jaccard_threshold}")
    print(f"{'='*60}\n")
    
    # Ambil data jaringan dari session
    network_data = request.session.get('glod_network_data')
    
    if not network_data:
        return render(request, 'glod_app/result.html', {
            'error': 'Data jaringan tidak ditemukan. Silakan ulangi dari awal.'
        })
    
    try:
        # Buat graph NetworkX dari data
        G = nx.Graph()
        
        # Tambahkan nodes
        for node in network_data['nodes']:
            G.add_node(node['id'])
        
        # Tambahkan edges
        for edge in network_data['edges']:
            G.add_edge(edge['source'], edge['target'])
        
        print(f"Graph created: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")
        
        # Jalankan algoritma GLOD
        glod = GLODAlgorithm(G, alpha=alpha, jaccard_threshold=jaccard_threshold)
        print(f"GLODAlgorithm initialized with alpha={glod.alpha}, jaccard_threshold={glod.jaccard_threshold}")
        
        communities, shen_mod, lazar_mod, nicosia_mod = glod.run(seed_value=42)  # Use fixed seed untuk hasil konsisten
        
        # ===== HITUNG NORMALIZED NODE CUT DAN CONDUCTANCE =====
        psi_scores = [glod.calculate_psi_normalized_node_cut(c) for c in communities]
        avg_psi = sum(psi_scores) / len(psi_scores) if psi_scores else 0.0
        
        conductance_scores = [glod.calculate_conductance(c) for c in communities]
        avg_conductance = sum(conductance_scores) / len(conductance_scores) if conductance_scores else 0.0
        
        print(f"Psi Node Cut scores: {psi_scores}")
        print(f"Average Psi: {avg_psi}")
        print(f"Conductance scores: {conductance_scores}")
        print(f"Average Conductance: {avg_conductance}")
        
        # Format hasil komunitas dengan overlap dan PSI
        community_results = []
        
        # Hitung nodes yang overlap (ada di multiple communities)
        node_community_count = {}
        for community in communities:
            for node in community:
                node_community_count[node] = node_community_count.get(node, 0) + 1
        
        # Buat dict untuk overlap per komunitas
        overlapping_nodes = set(node for node, count in node_community_count.items() if count > 1)
        
        for idx, community in enumerate(communities, 1):
            # Hitung PSI untuk komunitas ini
            psi_value = glod.calculate_psi_normalized_node_cut(community)
            
            # Hitung overlap untuk komunitas ini
            overlap_in_community = community & overlapping_nodes
            
            community_results.append({
                'id': idx,
                'size': len(community),
                'members': sorted(list(community)),
                'overlap_count': len(overlap_in_community),
                'overlap_members': sorted(list(overlap_in_community)),
                'psi': round(psi_value, 4)
            })
        
        # ===== HITUNG NMI METRICS (JIKA ADA GROUND TRUTH) =====
        nmi_metrics = {
            'nmi_lfk': None,
            'nmi_max': None,
            'rnmi': None
        }
        
        # Cek apakah ada ground truth di POST
        ground_truth_json = request.POST.get('ground_truth_data')
        if ground_truth_json:
            try:
                ground_truth_data = json.loads(ground_truth_json)
                # Convert ground truth data ke List[Set]
                ground_truth_comms = [set(comm) for comm in ground_truth_data if comm]
                
                if ground_truth_comms and len(ground_truth_comms) > 0:
                    # Hitung NMI metrics dengan seed konsisten
                    nmi_results = glod.calculate_onmi_metrics(communities, ground_truth_comms, seed_value=42)
                    nmi_metrics['nmi_lfk'] = round(nmi_results['nmi_lfk'], 4)
                    nmi_metrics['nmi_max'] = round(nmi_results['nmi_max'], 4)
                    nmi_metrics['rnmi'] = round(nmi_results['rnmi'], 4)
                    print(f"NMI metrics calculated successfully")
                else:
                    print(f"Ground truth data kosong atau tidak valid")
            except Exception as e:
                print(f"Warning: Gagal menghitung NMI metrics: {str(e)}")
        
        # Siapkan data untuk visualisasi
        # Assign warna untuk setiap komunitas
        node_colors = {}
        color_palette = [
            '#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8',
            '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52B788',
            '#FFB6C1', '#DDA15E', '#B4A7D6', '#A8DADC', '#E9C46A'
        ]
        
        for idx, community in enumerate(communities):
            color = color_palette[idx % len(color_palette)]
            for node in community:
                if node not in node_colors:
                    node_colors[node] = []
                node_colors[node].append({'community': idx + 1, 'color': color})
        
        # Format data untuk vis.js
        vis_nodes = []
        for node in network_data['nodes']:
            node_id = node['id']
            
            # Jika node ada di communities
            if node_id in node_colors:
                colors = node_colors[node_id]
                community_ids = [c['community'] for c in colors]
                title = f"{node_id}\nCommunities: {', '.join([str(cid) for cid in community_ids])}"
                node_color = colors[0]['color']
            else:
                # Node tidak ada di komunitas manapun (unlabeled)
                colors = [{'color': '#97C2FC'}]
                community_ids = []
                title = f"{node_id}\n(Unlabeled)"
                node_color = '#97C2FC'
            
            # Jika node ada di multiple communities, gunakan warna pertama
            vis_nodes.append({
                'id': node_id,
                'label': node['label'],
                'color': node_color,
                'title': title,
                'communities': community_ids
            })
        
        vis_edges = []
        for edge in network_data['edges']:
            vis_edges.append({
                'from': edge['source'],
                'to': edge['target'],
                'score': edge.get('score', 0)
            })
        
        context = {
            'num_communities': len(communities),
            'shen_modularity': round(shen_mod, 4),
            'lazar_modularity': round(lazar_mod, 4),
            'nicosia_modularity': round(nicosia_mod, 4),
            'avg_psi_node_cut': round(avg_psi, 4),
            'avg_conductance': round(avg_conductance, 4),
            'nmi_lfk': nmi_metrics['nmi_lfk'],
            'nmi_max': nmi_metrics['nmi_max'],
            'rnmi': nmi_metrics['rnmi'],
            'has_nmi_metrics': any(v is not None for v in nmi_metrics.values()),
            'alpha': alpha,
            'jaccard_threshold': jaccard_threshold,
            'communities': community_results,
            'communities_json': json.dumps(community_results),
            'total_nodes': G.number_of_nodes(),
            'total_edges': G.number_of_edges(),
            'vis_nodes': json.dumps(vis_nodes),
            'vis_edges': json.dumps(vis_edges),
            'color_palette': color_palette[:len(communities)]
        }
        
        return render(request, 'glod_app/result.html', context)
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        return render(request, 'glod_app/result.html', {
            'error': f'Error saat memproses algoritma GLOD: {str(e)}',
            'error_detail': error_detail
        })


@require_http_methods(["POST"])
@csrf_exempt
def download_community_data(request):
    """
    Download tabel hasil komunitas dalam format CSV atau XLSX
    """
    try:
        # Ambil data komunitas dari POST
        communities_json = request.POST.get('communities_json')
        format_type = request.POST.get('format', 'csv').lower()
        
        if not communities_json:
            return HttpResponse('Data komunitas tidak ditemukan', status=400)
        
        communities = json.loads(communities_json)
        
        if format_type == 'xlsx':
            return generate_xlsx(communities)
        else:
            return generate_csv(communities)
        
    except Exception as e:
        import traceback
        return HttpResponse(f'Error: {str(e)}\n{traceback.format_exc()}', status=500)


def generate_csv(communities):
    """Generate CSV file dari data komunitas"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="hasil_komunitas_glod_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Tulis BOM untuk Excel supaya bisa baca UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow([
        'Komunitas ID',
        'Jumlah Anggota',
        'Jumlah Overlap',
        'Anggota Protein',
        'Protein Overlap',
        'Normalized Node Cut (Ψ)'
    ])
    
    # Data dari tabel
    for community in communities:
        members_str = ', '.join(community.get('members', []))
        overlap_members_str = ', '.join(community.get('overlap_members', []))
        
        writer.writerow([
            f"Komunitas {community.get('id', '')}",
            community.get('size', ''),
            community.get('overlap_count', ''),
            members_str,
            overlap_members_str,
            community.get('psi', '')
        ])
    
    return response


def generate_xlsx(communities):
    """Generate XLSX file dari data komunitas"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hasil Komunitas'
    
    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Header
    headers = [
        'Komunitas ID',
        'Jumlah Anggota',
        'Jumlah Overlap',
        'Anggota Protein',
        'Protein Overlap',
        'Normalized Node Cut (Ψ)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    
    # Data
    for row_num, community in enumerate(communities, 2):
        members_str = ', '.join(community.get('members', []))
        overlap_members_str = ', '.join(community.get('overlap_members', []))
        
        data = [
            f"Komunitas {community.get('id', '')}",
            community.get('size', ''),
            community.get('overlap_count', ''),
            members_str,
            overlap_members_str,
            community.get('psi', '')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = center_alignment
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="hasil_komunitas_glod_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

